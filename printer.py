# -*- coding: utf-8 -*-
"""
Created on Tue Jan 18 15:25:47 2022

@author: yuki
"""
#https://qiita.com/komanaki/items/dc7d43b434a5612570f9

import os
import argparse

from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook


#色んなサイズのフォントの準備
font_bold_60 = ImageFont.truetype("fonts/MPLUS1p-Bold.ttf", 60)
font_bold_52 = ImageFont.truetype("fonts/MPLUS1p-Bold.ttf", 52)
font_bold_28 = ImageFont.truetype("fonts/MPLUS1p-Bold.ttf", 28)
font_medium_32 = ImageFont.truetype("fonts/MPLUS1p-Medium.ttf", 32)



#ファイルを開く
excel = load_workbook("notes.xlsx")
#一番目の表計算を使う
sheet = excel.active

#データを読んでおきます
exercice_name = sheet["B1"].value
exercice_date = sheet["D1"].value
exercice_lesson = sheet["F1"].value

#3行目から表計算を読んで、10列目まで
for row in sheet.iter_rows(min_row=3, max_col=10, values_only=True):
    #一つ行からデータを引き出す
    lastname, firstname, rank, syntax_rank, syntax_note, objective_rank, objective_note, quality_rank, quality_note, notes = row

    #ランクがない場合は多分、結果の打ち込みがまだ
    if rank is None:
        print(f"空きランク、スキップ...")
        continue

    print(f"--> {lastname} {firstname} ● ランク {rank}")

    #テンプレートのイメージを開く
    with Image.open("template.png") as im:
        draw = ImageDraw.Draw(im)

        # テキストの印刷
        draw.text((92, 80), f"{lastname}\n{firstname}", font=font_bold_52, fill=(0,0,0))
        #anchor="mm"の意味は、中央揃えで書く
        draw.text((1296, 196), exercice_name, font=font_bold_60, anchor="mm", fill=(255,255,255))
        draw.text((1298, 902), exercice_date.strftime("%d/%m/%Y"), font=font_bold_28, fill=(0,0,0))
        draw.text((1610, 902), exercice_lesson, font=font_bold_28, fill=(0,0,0))
        #ここからは複数行テキスト
        draw.multiline_text((500, 384), syntax_note or "", font=font_medium_32, anchor="lm", fill=(0,0,0))
        draw.multiline_text((500, 544), objective_note or "", font=font_medium_32, anchor="lm", fill=(0,0,0))
        draw.multiline_text((500, 704), quality_note or "", font=font_medium_32, anchor="lm", fill=(0,0,0))
        draw.multiline_text((280, 870), notes or "", font=font_medium_32, anchor="lm", fill=(255,255,255))

        #文字ランクのイメージのサイズが違っているので、
        #中央揃えのために、正しいx位置を書いておこう
        ranks_x_position = {
            "B": 1496,
            "A": 1466,
            "S": 1490,
            "SS": 1384,
            "SSS": 1278
        }

        #文字ランクのイメージの印刷
        with Image.open(f"rank_{rank}.png") as im_rank:
            im.paste(im_rank, (ranks_x_position[rank], 486), im_rank)

        #構文規則ランクのイメージの印刷
        with Image.open(f"rank_{syntax_rank}.png") as im_syntax_rank:
            im_syntax_rank.thumbnail((94, 94))
            im.paste(im_syntax_rank, (380 if syntax_rank != "A" else 375, 340), im_syntax_rank)

        #目的ランクのイメージの印刷
        with Image.open(f"rank_{objective_rank}.png") as im_objective_rank:
            im_objective_rank.thumbnail((94, 94))
            im.paste(im_objective_rank, (380 if objective_rank != "A" else 375, 500), im_objective_rank)

        #コード質のイメージの印刷
        with Image.open(f"rank_{quality_rank}.png") as im_quality_rank:
            im_quality_rank.thumbnail((94, 94))
            im.paste(im_quality_rank, (380 if quality_rank != "A" else 375, 660), im_quality_rank)

        #最後はイメージを保存する
        im.save(f"{firstname} {lastname} - {exercice_name}.png")